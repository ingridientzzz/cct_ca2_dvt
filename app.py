import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import openpyxl
import os

# Set Streamlit page layout to wide
st.set_page_config(layout='wide')

# Define color and font constants for consistent styling
FONT_FAMILY = '"Helvetica Neue", Helvetica, Arial, sans-serif'
COLOR_PRIMARY = '#228C8A'    # Vibrant Teal
COLOR_SECONDARY = '#F6C85F'  # Warm Gold
COLOR_MALE = '#6F4E7C'       # Deep Plum
COLOR_FEMALE = '#E17C72'     # Soft Coral
COLOR_FEMALE_2011_DIST = '#F2B5A9'  # Lighter Soft Coral

# --- FUNCTIONS FOR LOADING DATA ---
def parse_excel_file(excel_file):
    workbook = openpyxl.load_workbook(excel_file)

    tables_dfs = {
        f'{ws.title}_{tbl.name}': pd.DataFrame(
            [[cell.value for cell in row] for row in ws[tbl.ref][1:]],
            columns=[cell.value for cell in ws[tbl.ref][0]]
        ) for ws in workbook for tbl in ws.tables.values() if tbl.ref
    }

    return tables_dfs

@st.cache_data
def load_data(excel_file_path):
    try:
        tables_dfs = parse_excel_file(excel_file_path)
        df_density_raw = tables_dfs.get('MYE5_Table8', pd.DataFrame())
        df_age_gender_raw = tables_dfs.get('MYEB1_Table9', pd.DataFrame())
    except Exception as e:
        print(f"Error loading or parsing Excel file: {e}")
        df_density_raw = pd.DataFrame()
        df_age_gender_raw = pd.DataFrame()

    return df_age_gender_raw, df_density_raw

@st.cache_data
def preprocess_density_data_wide(df_density_raw):
    if df_density_raw.empty:
        return pd.DataFrame()

    df_density = df_density_raw.copy()

    # Rename columns
    rename_map = {
        'Code': 'code', 'Name': 'name', 'Geography': 'geography',
        'Area (sq km)': 'area_sq_km',
        'Estimated Population mid-2022': 'population_2022_abs',
        '2022 people per sq. km': 'density_2022',
        'Estimated Population mid-2011': 'population_2011_abs',
        '2011 people per sq. km': 'density_2011'
    }
    df_density = df_density.rename(columns={k: v for k, v in rename_map.items() if k in df_density.columns})

    # Convert density columns to numeric
    density_cols = ['density_2011', 'density_2022']
    for col in density_cols:
        if col in df_density.columns:
            df_density[col] = pd.to_numeric(df_density[col], errors='coerce')
        else:
            df_density[col] = pd.NA

    # Select final columns
    final_cols = ['code', 'name', 'geography', 'density_2011', 'density_2022', 'area_sq_km']
    existing_final_cols = [col for col in final_cols if col in df_density.columns]

    return df_density[existing_final_cols]

@st.cache_data
def melt_density_data(df_density_wide):
    if df_density_wide.empty:
        return pd.DataFrame()

    needed = ['code', 'name', 'geography', 'density_2011', 'density_2022']
    if not all(col in df_density_wide.columns for col in needed):
        return pd.DataFrame()

    df_density_melted = df_density_wide.melt(
        id_vars=['code', 'name', 'geography'],
        value_vars=['density_2011', 'density_2022'],
        var_name='Year_Col_Density',
        value_name='Density'
    )
    df_density_melted['Year'] = df_density_melted['Year_Col_Density'].str.extract(r'(\d+)').astype(int)
    df_density_melted = df_density_melted.drop(columns=['Year_Col_Density'])

    return df_density_melted[['code', 'name', 'geography', 'Year', 'Density']].dropna(subset=['Density'])

@st.cache_data
def preprocess_age_gender_data(df_age_gender_raw, bins, labels):
    if df_age_gender_raw.empty:
        return pd.DataFrame(), pd.DataFrame()

    df_age_gender_detail = df_age_gender_raw.copy()

    # Convert age to numeric
    df_age_gender_detail['age_numeric'] = df_age_gender_detail['age'].astype(str).replace('90+', '90').astype(int)

    # Create age bands
    df_age_gender_detail['age_band'] = pd.cut(
        df_age_gender_detail['age_numeric'],
        bins=bins, labels=labels, right=True, include_lowest=True
    )

    # Melt data for year comparison
    df_age_gender_melted = df_age_gender_detail.melt(
        id_vars=['code', 'name', 'geography', 'sex', 'age', 'age_numeric', 'age_band'],
        value_vars=['population_2011', 'population_2022'],
        var_name='Year_Col', value_name='Population'
    )

    # Extract year and convert to numeric
    df_age_gender_melted['Year'] = df_age_gender_melted['Year_Col'].str.extract(r'(\d+)').astype(int)
    df_age_gender_melted = df_age_gender_melted.drop(columns=['Year_Col'])

    # Convert population to numeric
    df_age_gender_melted['Population'] = pd.to_numeric(df_age_gender_melted['Population'], errors='coerce').fillna(0)
    df_age_gender_detail['population_2011'] = pd.to_numeric(df_age_gender_detail['population_2011'], errors='coerce').fillna(0)
    df_age_gender_detail['population_2022'] = pd.to_numeric(df_age_gender_detail['population_2022'], errors='coerce').fillna(0)

    # Sort data
    df_age_gender_melted = df_age_gender_melted.sort_values(by=['name', 'Year', 'sex', 'age_numeric'])

    return df_age_gender_detail, df_age_gender_melted

@st.cache_data
def create_merged_df(df_age_gender_melted, df_density_melted):
    if df_age_gender_melted.empty or df_density_melted.empty:
        return pd.DataFrame()

    df_age_gender_melted['Year'] = df_age_gender_melted['Year'].astype(int)
    df_density_melted['Year'] = df_density_melted['Year'].astype(int)

    merged_df = pd.merge(
        df_age_gender_melted,
        df_density_melted,
        on=['code', 'name', 'geography', 'Year'],
        how='left'
    )

    return merged_df

def create_empty_figure_graph(title_text):
    """
    Create an empty Plotly figure with title and hidden axes.
    """
    fig = go.Figure()
    fig.update_layout(
        title=title_text,
        xaxis={'visible': False},
        yaxis={'visible': False},
        font={'family': FONT_FAMILY}
    )
    return fig

def get_vrect_coords_from_age_band(age_band_str):
    """Returns x0, x1 coordinates for age band highlighting in charts"""
    if age_band_str == 'All Ages':
        return None, None
    if age_band_str == '75+':
        return '75', '90+'
    parts = age_band_str.split('-')
    if len(parts) == 2:
        return str(parts[0]), str(parts[1])
    return None, None

# Define mappings and options for filters and age bands
genders_map = {'M': 'Male', 'F': 'Female'}
bins = [-1, 17, 24, 39, 59, 74, np.inf]
labels = ['0-17', '18-24', '25-39', '40-59', '60-74', '75+']
age_bands_options = ['All Ages'] + labels

# --- LOAD DATA ---
data_folder = os.path.join(os.path.dirname(__file__), 'data')
excel_file_path = os.path.join(data_folder, 'mye22final.xlsx')

# Load raw data
df_age_gender_raw, df_density_raw = load_data(excel_file_path)

# Process data
df_density_wide = preprocess_density_data_wide(df_density_raw)
df_density_melted = melt_density_data(df_density_wide)
df_age_gender_detail, df_age_gender_melted = preprocess_age_gender_data(df_age_gender_raw, bins, labels)
merged_df = create_merged_df(df_age_gender_melted, df_density_melted)

# Prepare filter options
all_locations = []
all_geographies = []

if not df_age_gender_melted.empty and 'name' in df_age_gender_melted.columns:
    all_locations.extend(df_age_gender_melted['name'].unique())

if not df_age_gender_melted.empty and 'geography' in df_age_gender_melted.columns:
    all_geographies.extend(df_age_gender_melted['geography'].unique())

if not df_density_wide.empty and 'name' in df_density_wide.columns:
    all_locations.extend(df_density_wide['name'].unique())

if not df_density_wide.empty and 'geography' in df_density_wide.columns:
    all_geographies.extend(df_density_wide['geography'].unique())

all_locations = sorted(list(set(all_locations)))
all_geographies = sorted(list(set(all_geographies)))

# --- SIDEBAR FILTERS ---
st.sidebar.header("Global Filters")

# Geography filter
selected_geographies = st.sidebar.multiselect(
    "Select Geography Type(s):",
    options=all_geographies,
    default=all_geographies
)

# Filter locations based on selected geographies
locations_for_filter = []
if selected_geographies:
    if not df_age_gender_melted.empty and 'geography' in df_age_gender_melted.columns:
        locations_for_filter.extend(
            df_age_gender_melted[df_age_gender_melted['geography'].isin(selected_geographies)]['name'].unique()
        )
    if not df_density_wide.empty and 'geography' in df_density_wide.columns:
        locations_for_filter.extend(
            df_density_wide[df_density_wide['geography'].isin(selected_geographies)]['name'].unique()
        )
    locations_for_filter = sorted(list(set(locations_for_filter)))
else:
    locations_for_filter = all_locations

# Default locations for filter
default_locations = [loc for loc in ['ENGLAND', 'SCOTLAND', 'WALES', 'NORTHERN IRELAND']
                    if loc in locations_for_filter]
if not default_locations and locations_for_filter:
    default_locations = locations_for_filter[:min(4, len(locations_for_filter))]

# Location filter
selected_locations = st.sidebar.multiselect(
    "Select Location(s):",
    options=locations_for_filter,
    default=default_locations
)

# --- FILTER DATAFRAMES ---
# Filter age/gender data
df_display_melted = pd.DataFrame()
df_display_detail = pd.DataFrame()
df_display_density = pd.DataFrame()
df_display_merged = pd.DataFrame()

if not df_age_gender_melted.empty and 'name' in df_age_gender_melted.columns and 'geography' in df_age_gender_melted.columns:
    df_display_melted = df_age_gender_melted[
        df_age_gender_melted['name'].isin(selected_locations) &
        df_age_gender_melted['geography'].isin(selected_geographies)
    ]

if not df_age_gender_detail.empty and 'name' in df_age_gender_detail.columns and 'geography' in df_age_gender_detail.columns:
    df_display_detail = df_age_gender_detail[
        df_age_gender_detail['name'].isin(selected_locations) &
        df_age_gender_detail['geography'].isin(selected_geographies)
    ]

if not df_density_wide.empty and 'name' in df_density_wide.columns and 'geography' in df_density_wide.columns:
    df_display_density = df_density_wide[
        df_density_wide['name'].isin(selected_locations) &
        df_density_wide['geography'].isin(selected_geographies)
    ]

if not merged_df.empty and 'name' in merged_df.columns and 'geography' in merged_df.columns:
    df_display_merged = merged_df[
        merged_df['name'].isin(selected_locations) &
        merged_df['geography'].isin(selected_geographies)
    ]

# --- YEAR, SEX, AGE BAND FILTERS ---
# Year filter
global_selected_year_display = st.sidebar.selectbox(
    "Select Year(s) to Display:",
    options=["Comparison (2011 & 2022)", "2011 Only", "2022 Only"],
    index=0
)

# Sex filter
sex_options_map = {"Both": ["M", "F"], "Male": ["M"], "Female": ["F"]}
global_selected_gender_display_label = st.sidebar.selectbox(
    "Select Sex (for applicable charts):",
    options=["Both", "Male", "Female"],
    index=0
)
global_selected_sex_codes = sex_options_map[global_selected_gender_display_label]

# Age band filter
global_selected_age_band = st.sidebar.selectbox(
    "Select Age Band (for applicable charts):",
    options=age_bands_options,
    index=0
)

# --- MAIN PAGE CONTENT ---
st.title("UK Population Dashboard: 2011 vs 2022")
st.markdown("---")

# 1. Population Density Comparison
st.header("1. Population Density Comparison")
if df_display_density.empty:
    st.plotly_chart(create_empty_figure_graph("No density data available for selected filters."), use_container_width=True)
else:
    # Plot density bars
    filtered_df_density = df_display_density.sort_values('name')
    fig_density = go.Figure()
    plot_data_exists = False

    # Add 2011 data if selected
    if global_selected_year_display in ["2011 Only", "Comparison (2011 & 2022)"]:
        if 'density_2011' in filtered_df_density.columns:
            fig_density.add_trace(go.Bar(
                x=filtered_df_density['name'],
                y=filtered_df_density['density_2011'],
                name='Density 2011',
                marker_color=COLOR_SECONDARY,
                hovertemplate="**%{x}**<br>2011 Density: %{y:.1f} per sq km<extra></extra>"
            ))
            plot_data_exists = True

    # Add 2022 data if selected
    if global_selected_year_display in ["2022 Only", "Comparison (2011 & 2022)"]:
        if 'density_2022' in filtered_df_density.columns:
            fig_density.add_trace(go.Bar(
                x=filtered_df_density['name'],
                y=filtered_df_density['density_2022'],
                name='Density 2022',
                marker_color=COLOR_PRIMARY,
                hovertemplate="**%{x}**<br>2022 Density: %{y:.1f} per sq km<extra></extra>"
            ))
            plot_data_exists = True

    # Set chart title based on year selection
    title_density = 'Population Density'
    if global_selected_year_display != "Comparison (2011 & 2022)":
        title_density += f' ({global_selected_year_display.replace(" Only", "")})'
    else:
        title_density += ': 2011 vs 2022'

    # Update layout
    if plot_data_exists:
        fig_density.update_layout(
            title=title_density,
            xaxis_title='Location',
            yaxis_title='People per Square Kilometer',
            barmode='group' if global_selected_year_display == "Comparison (2011 & 2022)" else 'overlay',
            hovermode='x unified',
            legend_title_text='Year',
            font={'family': FONT_FAMILY},
            margin={'l': 40, 'r': 20, 't': 60, 'b': 40}
        )
        st.plotly_chart(fig_density, use_container_width=True)
    else:
        st.plotly_chart(create_empty_figure_graph("No density data for selected year(s)."), use_container_width=True)

st.markdown("---")

# 2. Population Age Distribution
st.header("2. Population Age Distribution")

# Select location for age distribution
loc_for_age_dist = None
if not df_display_detail.empty and 'name' in df_display_detail.columns:
    if selected_locations and selected_locations[0] in df_display_detail['name'].unique():
        loc_for_age_dist = selected_locations[0]
        if len(selected_locations) > 1:
            st.info(f"Displaying age distribution for {loc_for_age_dist} (first selected).")
    elif len(df_display_detail['name'].unique()) > 0:
        loc_for_age_dist = df_display_detail['name'].unique()[0]
        st.info(f"Displaying age distribution for {loc_for_age_dist} (first available in filtered data).")

if loc_for_age_dist and not df_display_detail.empty:
    # Create age distribution chart
    fig_age_gender = go.Figure()
    data_found = False

    # Filter data for selected location and sex
    if all(col in df_display_detail.columns for col in ['name', 'sex', 'age_numeric']):
        chart_df_age = df_display_detail[
            (df_display_detail['name'] == loc_for_age_dist) &
            (df_display_detail['sex'].isin(global_selected_sex_codes))
        ].sort_values('age_numeric')

        # Add traces for each gender
        for gender_code in global_selected_sex_codes:
            gender_df = chart_df_age[chart_df_age['sex'] == gender_code]
            if not gender_df.empty:
                gender_label = genders_map[gender_code]
                color_2022 = COLOR_MALE if gender_code == 'M' else COLOR_FEMALE
                color_2011 = COLOR_PRIMARY if gender_code == 'M' else COLOR_FEMALE_2011_DIST

                # Add 2011 data if selected
                if global_selected_year_display in ["2011 Only", "Comparison (2011 & 2022)"]:
                    if 'population_2011' in gender_df.columns:
                        fig_age_gender.add_trace(go.Scatter(
                            x=gender_df['age'],
                            y=gender_df['population_2011'],
                            name=f'{gender_label} 2011',
                            mode='lines',
                            marker_color=color_2011,
                            line=dict(width=2, dash='dash'),
                            hovertemplate=f"<b>{gender_label} - Age: %{{x}}</b><br>2011 Pop: %{{y:,}}<extra></extra>"
                        ))
                        data_found = True

                # Add 2022 data if selected
                if global_selected_year_display in ["2022 Only", "Comparison (2011 & 2022)"]:
                    if 'population_2022' in gender_df.columns:
                        fig_age_gender.add_trace(go.Scatter(
                            x=gender_df['age'],
                            y=gender_df['population_2022'],
                            name=f'{gender_label} 2022',
                            mode='lines',
                            marker_color=color_2022,
                            line=dict(width=2),
                            hovertemplate=f"<b>{gender_label} - Age: %{{x}}</b><br>2022 Pop: %{{y:,}}<extra></extra>"
                        ))
                        data_found = True

        if data_found:
            # Set chart title
            title_age_dist = f'Pop Age Distribution in {loc_for_age_dist}'
            if global_selected_year_display != "Comparison (2011 & 2022)":
                title_age_dist += f' ({global_selected_year_display.replace(" Only", "")})'

            # Update layout
            fig_age_gender.update_layout(
                title=title_age_dist,
                xaxis_title='Age',
                yaxis_title='Est. Population',
                xaxis={'type': 'category'},
                hovermode='x unified',
                legend_title_text='Sex & Year',
                font={'family': FONT_FAMILY},
                margin={'l': 40, 'r': 20, 't': 60, 'b': 40}
            )

            # Add age band highlight if selected
            if global_selected_age_band != 'All Ages':
                x0, x1 = get_vrect_coords_from_age_band(global_selected_age_band)
                if x0 and x1:
                    fig_age_gender.add_vrect(
                        x0=x0, x1=x1,
                        fillcolor="rgba(128,128,128,0.2)",
                        layer="below",
                        line_width=0
                    )

            st.plotly_chart(fig_age_gender, use_container_width=True)
        else:
            st.plotly_chart(create_empty_figure_graph(f"No age data to plot for {loc_for_age_dist} with current filters."), use_container_width=True)
    else:
        st.plotly_chart(create_empty_figure_graph("Required columns missing in age data."), use_container_width=True)
else:
    st.plotly_chart(create_empty_figure_graph("No age data available for selected filters."), use_container_width=True)

st.markdown("---")

# 3. Gender Population Comparison
st.header("3. Gender Population Comparison")

# Determine year for gender comparison
year_for_gender_comp = None
if global_selected_year_display == "2011 Only":
    year_for_gender_comp = 2011
elif global_selected_year_display == "2022 Only":
    year_for_gender_comp = 2022
elif global_selected_year_display == "Comparison (2011 & 2022)":
    year_for_gender_comp = 2022
    st.info("This chart defaults to 2022 data for 'Comparison (2011 & 2022)' year selection.")

if df_display_melted.empty or 'Year' not in df_display_melted.columns:
    st.plotly_chart(create_empty_figure_graph("No gender data available for selected filters."), use_container_width=True)
elif not year_for_gender_comp:
    st.plotly_chart(create_empty_figure_graph("Select a year to display gender comparison."), use_container_width=True)
else:
    # Filter data for selected year and age band
    chart_df_gender = df_display_melted[df_display_melted['Year'] == year_for_gender_comp].copy()

    if global_selected_age_band != 'All Ages' and 'age_band' in chart_df_gender.columns:
        chart_df_gender = chart_df_gender[chart_df_gender['age_band'] == global_selected_age_band]

    if 'sex' in chart_df_gender.columns:
        chart_df_gender = chart_df_gender[chart_df_gender['sex'].isin(global_selected_sex_codes)]

    if chart_df_gender.empty:
        st.plotly_chart(create_empty_figure_graph(
            f"No data for gender comparison in {year_for_gender_comp} with current filters."
        ), use_container_width=True)
    else:
        # Group data by location and sex
        if all(col in chart_df_gender.columns for col in ['name', 'sex', 'Population']):
            grouped_gender = chart_df_gender.groupby(['name', 'sex'])['Population'].sum().reset_index().sort_values(['name', 'sex'])

            # Create chart
            fig_gender = go.Figure()
            plot_data_exists = False

            # Add bars for each sex
            for sex_code in global_selected_sex_codes:
                gender_label = genders_map[sex_code]
                df_sex = grouped_gender[grouped_gender['sex'] == sex_code]

                if not df_sex.empty:
                    color = COLOR_MALE if sex_code == 'M' else COLOR_FEMALE
                    fig_gender.add_trace(go.Bar(
                        x=df_sex['name'],
                        y=df_sex['Population'],
                        name=gender_label,
                        marker_color=color,
                        hovertemplate=(
                            f"<b>%{{x}}</b><br>{gender_label}s: %{{y:,}}"
                            f"<br>Year: {year_for_gender_comp}"
                            f"<br>Age Band: {global_selected_age_band}<extra></extra>"
                        )
                    ))
                    plot_data_exists = True

            if plot_data_exists:
                # Update layout
                age_title_part = f"({global_selected_age_band})" if global_selected_age_band != 'All Ages' else ""
                fig_gender.update_layout(
                    title=f'Pop by Sex in {year_for_gender_comp} {age_title_part}',
                    xaxis_title='Location',
                    yaxis_title='Population',
                    barmode='group',
                    hovermode='x unified',
                    legend_title_text='Sex',
                    font={'family': FONT_FAMILY},
                    margin={'l': 40, 'r': 20, 't': 60, 'b': 40}
                )
                st.plotly_chart(fig_gender, use_container_width=True)
            else:
                st.plotly_chart(create_empty_figure_graph(
                    f"No data for selected sex(es) in {year_for_gender_comp} after grouping."
                ), use_container_width=True)
        else:
            st.plotly_chart(create_empty_figure_graph("Required columns missing for gender comparison."), use_container_width=True)

st.markdown("---")

# 4. Sex Ratio by Population Density
st.header("4. Sex Ratio by Population Density")
st.info("Sex ratio is Males per 100 Females. This chart ignores the global 'Select Sex' filter.")

if df_display_merged.empty or 'Density' not in df_display_merged.columns:
    st.plotly_chart(create_empty_figure_graph("No density data available for sex ratio analysis."), use_container_width=True)
else:
    # Determine years to plot
    years_to_plot = []
    if global_selected_year_display == "2011 Only":
        years_to_plot = [2011]
    elif global_selected_year_display == "2022 Only":
        years_to_plot = [2022]
    elif global_selected_year_display == "Comparison (2011 & 2022)":
        years_to_plot = [2011, 2022]

    if not years_to_plot:
        st.warning("Please select a year or 'Comparison' to display sex ratio by density.")
    else:
        # Filter data for selected years
        sex_ratio_data = df_display_merged[df_display_merged['Year'].isin(years_to_plot)]

        # Filter by age band if selected
        if global_selected_age_band != 'All Ages' and 'age_band' in sex_ratio_data.columns:
            sex_ratio_data = sex_ratio_data[sex_ratio_data['age_band'] == global_selected_age_band]

        if sex_ratio_data.empty:
            st.plotly_chart(create_empty_figure_graph(
                f"No data for sex ratio analysis with current filters."
            ), use_container_width=True)
        else:
            # Check required columns
            required_cols = ['code', 'name', 'geography', 'Year', 'Density', 'sex', 'Population']
            if all(col in sex_ratio_data.columns for col in required_cols):
                # Group by location, year, and sex
                pop_by_sex = sex_ratio_data.groupby(
                    ['code', 'name', 'geography', 'Year', 'Density', 'sex']
                )['Population'].sum().unstack(level='sex', fill_value=0).reset_index()

                # Ensure M/F columns exist
                if 'M' not in pop_by_sex.columns: pop_by_sex['M'] = 0
                if 'F' not in pop_by_sex.columns: pop_by_sex['F'] = 0

                # Calculate sex ratio
                pop_by_sex['sex_ratio'] = (pop_by_sex['M'] / pop_by_sex['F'].replace(0, np.nan)) * 100
                final_data = pop_by_sex.dropna(subset=['Density', 'sex_ratio'])

                if final_data.empty:
                    st.plotly_chart(create_empty_figure_graph(
                        f"Not enough data to calculate sex ratios for the scatter plot."
                    ), use_container_width=True)
                else:
                    # Create scatter plot
                    fig_sex_ratio = go.Figure()
                    plot_data_exists = False

                    # Add traces for each year
                    for year in years_to_plot:
                        year_data = final_data[final_data['Year'] == year]
                        if not year_data.empty:
                            marker_color = COLOR_PRIMARY if year == 2022 else COLOR_SECONDARY

                            # Ensure customdata columns exist
                            custom_data_cols = ['Year', 'M', 'F']
                            if all(col in year_data.columns for col in custom_data_cols):
                                fig_sex_ratio.add_trace(go.Scatter(
                                    x=year_data['Density'],
                                    y=year_data['sex_ratio'],
                                    mode='markers',
                                    name=f'Sex Ratio {year}',
                                    marker=dict(color=marker_color, size=8, opacity=0.7),
                                    text=year_data['name'],
                                    hovertemplate=(
                                        "<b>%{text} (%{customdata[0]})</b><br>"
                                        "Density: %{x:.1f} per sq km<br>"
                                        "Sex Ratio: %{y:.1f} (Males per 100 Females)<br>"
                                        "Males: %{customdata[1]:,}<br>Females: %{customdata[2]:,}"
                                        "<extra></extra>"
                                    ),
                                    customdata=year_data[custom_data_cols]
                                ))
                                plot_data_exists = True

                    if plot_data_exists:
                        # Update layout
                        age_band_title = f"(Age Band: {global_selected_age_band})" if global_selected_age_band != 'All Ages' else ""
                        title = f'Sex Ratio vs. Population Density {age_band_title}'
                        if len(years_to_plot) == 1:
                            title += f' - {years_to_plot[0]}'

                        fig_sex_ratio.update_layout(
                            title=title,
                            xaxis_title='Population Density (people per sq km)',
                            yaxis_title='Sex Ratio (Males per 100 Females)',
                            font={'family': FONT_FAMILY},
                            hovermode='closest',
                            legend_title_text='Year',
                            margin={'l': 40, 'r': 20, 't': 60, 'b': 40}
                        )
                        st.plotly_chart(fig_sex_ratio, use_container_width=True)
                    else:
                        st.plotly_chart(create_empty_figure_graph(
                            f"No valid sex ratio data to plot for selected year(s)."
                        ), use_container_width=True)
            else:
                st.plotly_chart(create_empty_figure_graph("Required columns missing for sex ratio analysis."), use_container_width=True)

st.markdown("---")
# footer
url = "https://github.com/ingridientzzz"
st.markdown(f"[:coffee: ingridientzzz]({url})")
